import os
import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction
from apps.catalogo.models import Material, Categoria
from apps.activos.models import Activo
from apps.logistica.models import Almacen

class Command(BaseCommand):
    help = 'Carga masiva de Materiales y Activos desde un archivo Excel (.xlsx)'

    def add_arguments(self, parser):
        parser.add_argument('excel_file', type=str, help='Ruta absoluta o relativa al archivo Excel')

    def handle(self, *args, **options):
        file_path = options['excel_file']

        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'El archivo no existe: {file_path}'))
            return

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Error al abrir Excel: {e}'))
            return

        # Ejecutamos todo en una transacción para integridad de datos
        with transaction.atomic():
            if 'Materiales' in wb.sheetnames:
                self.importar_materiales(wb['Materiales'])
            else:
                self.stdout.write(self.style.WARNING('No se encontró la hoja "Materiales". Saltando...'))

            if 'Activos' in wb.sheetnames:
                self.importar_activos(wb['Activos'])
            else:
                self.stdout.write(self.style.WARNING('No se encontró la hoja "Activos". Saltando...'))

    def importar_materiales(self, sheet):
        self.stdout.write(self.style.MIGRATE_HEADING('--- Importando Materiales ---'))
        count = 0
        # Estructura esperada (Fila 1 Cabeceras):
        # A: CODIGO | B: DESCRIPCION | C: UNIDAD | D: CATEGORIA | E: TIPO (CONSUMIBLE/ACTIVO_FIJO/EPP)
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Desempaquetar con seguridad (rellenar con None si faltan columnas)
            row_data = list(row) + [None] * (5 - len(row))
            codigo, descripcion, unidad, cat_nombre, tipo = row_data[:5]
            
            if not codigo: continue

            # 1. Buscar o Crear Categoría
            categoria = None
            if cat_nombre:
                nombre_cat = str(cat_nombre).strip().upper()
                # 1. Buscamos por nombre exacto
                categoria = Categoria.objects.filter(nombre=nombre_cat).first()
                
                if not categoria:
                    # 2. Si no existe por nombre, buscamos por CÓDIGO (Inferencia)
                    base_code = nombre_cat[:3].upper()
                    categoria = Categoria.objects.filter(codigo=base_code).first()
                    
                    if not categoria:
                        # 3. Solo si no existe ni nombre ni código, CREAMOS
                        categoria = Categoria.objects.create(nombre=nombre_cat, codigo=base_code)

            # 2. Crear o Actualizar Material
            material, created = Material.objects.update_or_create(
                codigo=str(codigo).strip().upper(),
                defaults={
                    'descripcion': str(descripcion).strip().upper() if descripcion else 'SIN DESCRIPCION',
                    'unidad_medida': str(unidad).strip().upper() if unidad else 'UND',
                    'categoria': categoria,
                    'tipo': str(tipo).strip().upper() if tipo else 'CONSUMIBLE',
                    'activo': True
                }
            )
            
            accion = "Creado" if created else "Actualizado"
            self.stdout.write(f"Material {codigo}: {accion}")
            count += 1
        
        self.stdout.write(self.style.SUCCESS(f'Total Materiales procesados: {count}'))

    def importar_activos(self, sheet):
        self.stdout.write(self.style.MIGRATE_HEADING('--- Importando Activos Fijos ---'))
        count = 0
        # Estructura esperada (Fila 1 Cabeceras):
        # A: CODIGO_ACTIVO | B: SERIE | C: NOMBRE | D: MARCA | E: MODELO | F: COD_MATERIAL_CATALOGO | G: ALMACEN_UBICACION
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = list(row) + [None] * (7 - len(row))
            cod_activo, serie, nombre, marca, modelo, cod_material, nombre_almacen = row_data[:7]

            if not cod_activo: continue

            # 1. Vincular con Material del Catálogo
            material = None
            if cod_material:
                material = Material.objects.filter(codigo=str(cod_material).strip().upper()).first()

            # 2. Buscar Almacén Inicial
            ubicacion = None
            if nombre_almacen:
                ubicacion = Almacen.objects.filter(nombre__icontains=str(nombre_almacen).strip()).first()

            # 3. Crear Activo
            activo, created = Activo.objects.update_or_create(
                codigo=str(cod_activo).strip().upper(),
                defaults={
                    'serie': str(serie).strip().upper() if serie else '',
                    'nombre': str(nombre).strip().upper() if nombre else 'ACTIVO SIN NOMBRE',
                    'marca': str(marca).strip().upper() if marca else '',
                    'modelo': str(modelo).strip().upper() if modelo else '',
                    'material': material,
                    'ubicacion': ubicacion,
                    'estado': 'DISPONIBLE',
                }
            )
            if created: count += 1
        
        self.stdout.write(self.style.SUCCESS(f'Total Activos creados: {count}'))