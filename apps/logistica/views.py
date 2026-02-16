from django.shortcuts import render, redirect, get_object_or_404
from django.db import transaction
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.http import HttpResponse, JsonResponse
from django.template.loader import get_template
from django.db.models import Prefetch, Q, F, Case, When, Value, DecimalField, Window, Sum
from django.db.models.functions import Coalesce
from django.utils import timezone
from decimal import Decimal
from xhtml2pdf import pisa
import qrcode
from io import BytesIO
import base64
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import json
from django.urls import reverse

# Importamos modelos y formularios locales
from .models import Movimiento, DetalleMovimiento, Stock, Almacen, Material, Proyecto, Requerimiento, Existencia, DetalleRequerimiento
from .forms import MovimientoForm, DetalleMovimientoFormSet, RequerimientoForm, DetalleRequerimientoFormSet, ImportarDatosForm
from .services import KardexService
from apps.rrhh.models import Trabajador
from apps.activos.models import Activo, AsignacionActivo, Kit
from apps.catalogo.models import Categoria # Necesario para crear categorías al vuelo
from apps.core.models import Configuracion

# ==========================================
# 1. REPORTES Y PDF
# ==========================================

# --- NUEVA VISTA: CAMBIAR ALMACÉN ACTIVO ---
def cambiar_almacen_sesion(request, almacen_id):
    """
    Establece el almacén activo en la sesión del usuario.
    """
    almacen = get_object_or_404(Almacen, id=almacen_id)
    request.session['almacen_activo_id'] = str(almacen.id)
    messages.success(request, f"Trabajando ahora en: {almacen.nombre}")
    
    # Redirigir a la página donde estaba (o al dashboard si no hay referer)
    return redirect(request.META.get('HTTP_REFERER', 'inventario_list'))

def limpiar_almacen_sesion(request):
    """
    Elimina el almacén activo de la sesión (Modo Vista Global).
    """
    if 'almacen_activo_id' in request.session:
        del request.session['almacen_activo_id']
        messages.info(request, "Modo Vista Global activado.")
    
    return redirect(request.META.get('HTTP_REFERER', 'dashboard'))

def generar_vale_pdf(request, movimiento_id):
    movimiento = get_object_or_404(Movimiento, id=movimiento_id)
    config = Configuracion.objects.first()
    
    # Generar Código QR con datos clave
    qr_data = f"DOC: {movimiento.nota_ingreso or 'S/N'}\nFECHA: {movimiento.fecha.strftime('%d/%m/%Y')}\nREF: {movimiento.id}"
    qr = qrcode.make(qr_data)
    buffer = BytesIO()
    qr.save(buffer, format="PNG")
    qr_img = base64.b64encode(buffer.getvalue()).decode()

    # Optimizamos la consulta y preparamos los detalles
    detalles = movimiento.detalles.select_related('material', 'requerimiento').all()

    # Lógica para columna dinámica:
    # Mostrar columna si NO hay requerimiento global Y al menos un ítem tiene requerimiento específico
    mostrar_columna_req = False
    if not movimiento.requerimiento and any(d.requerimiento for d in detalles):
        mostrar_columna_req = True

    template_path = 'logistica/vale_pdf.html'
    context = {
        'movimiento': movimiento,
        'detalles': detalles,
        'config': config,
        'qr_code': qr_img,
        'titulo': f"VALE DE {movimiento.get_tipo_display().upper()}",
        'mostrar_columna_req': mostrar_columna_req,
    }

    response = HttpResponse(content_type='application/pdf')
    nombre_archivo = movimiento.nota_ingreso if movimiento.nota_ingreso else movimiento.codigo_visual()
    filename = f"Vale_{nombre_archivo}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'

    template = get_template(template_path)
    html = template.render(context)

    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
       return HttpResponse('Error al generar PDF <pre>' + html + '</pre>')
    return response

# ==========================================
# 2. VISTAS DE INVENTARIO Y LISTADOS
# ==========================================

def inventario_list(request):
    """
    Reporte de Stock Físico con Búsqueda Inteligente.
    """
    query = request.GET.get('q')
    filtro = request.GET.get('filtro') # Nuevo parámetro para filtrar alertas

    stocks_filter = Stock.objects.select_related('material')

    if query:
        stocks_filter = stocks_filter.filter(
            Q(material__codigo__icontains=query) |
            Q(material__descripcion__icontains=query)
        )
    
    # FILTRO POR CONTEXTO DE ALMACÉN
    almacen_activo = getattr(request, 'almacen_activo', None)
    if almacen_activo:
        stocks_filter = stocks_filter.filter(almacen=almacen_activo)

    # Si viene del dashboard (clic en tarjeta roja), filtramos solo los críticos
    if filtro == 'critico':
        stocks_filter = stocks_filter.filter(cantidad__lte=F('cantidad_minima'), cantidad_minima__gt=0)
    elif filtro == 'advertencia':
        stocks_filter = stocks_filter.filter(
            cantidad__gt=F('cantidad_minima'),
            cantidad__lte=F('cantidad_minima') * Decimal('1.2'),
            cantidad_minima__gt=0
        )

    # Si hay almacén activo, solo mostramos ese almacén en la lista
    if almacen_activo:
        # CORRECCIÓN: Usamos filter(id=...) para mantener el prefetch_related con el filtro de stocks (búsqueda/alertas)
        almacenes = Almacen.objects.filter(id=almacen_activo.id).prefetch_related(
            Prefetch('stocks', queryset=stocks_filter)
        ).distinct()
    else:
        almacenes = Almacen.objects.prefetch_related(
            Prefetch('stocks', queryset=stocks_filter)
        ).distinct()

    context = {
        'almacenes': almacenes,
        'busqueda': query
    }
    return render(request, 'logistica/inventario.html', context)

def movimiento_list(request):
    """
    Pantalla Principal de Movimientos.
    Permite filtrar por estado (ej: ?estado=BORRADOR)
    """
    estado = request.GET.get('estado')
    
    qs = Movimiento.objects.select_related(
        'almacen_origen', 
        'almacen_destino', 
        'torre_destino',
        'proyecto'
    ).order_by('-fecha')

    # FILTRO POR CONTEXTO DE ALMACÉN
    almacen_activo = getattr(request, 'almacen_activo', None)
    if almacen_activo:
        qs = qs.filter(
            Q(almacen_origen=almacen_activo) | 
            Q(almacen_destino=almacen_activo)
        )

    if estado:
        qs = qs.filter(estado=estado)
    else:
        qs = qs[:50] # Filtramos los últimos 50 para no saturar si no hay filtro
    
    context = {
        'movimientos': qs,
        'estado_filtro': estado
    }
    return render(request, 'logistica/movimiento_list.html', context)

def kardex_producto(request, almacen_id, material_id):
    """
    Muestra la historia clínica de un material específico en un almacén.
    """
    almacen = get_object_or_404(Almacen, id=almacen_id)
    material = get_object_or_404(Material, id=material_id)

    # 1. Consulta Optimizada con Window Functions
    # Calculamos el saldo acumulado y clasificamos E/S directamente en la BD
    movimientos = DetalleMovimiento.objects.filter(
        material=material,
        movimiento__estado='CONFIRMADO'
    ).filter(
        Q(movimiento__almacen_origen=almacen) | 
        Q(movimiento__almacen_destino=almacen)
    ).annotate(
        # Determinar flujo (+/-) respecto al almacén actual
        flujo_cantidad=Case(
            When(movimiento__almacen_destino=almacen, then=F('cantidad')), # Es Ingreso (+)
            When(movimiento__almacen_origen=almacen, then=F('cantidad') * -1), # Es Salida (-)
            default=Value(0),
            output_field=DecimalField(max_digits=12, decimal_places=2)
        ),
        # Clasificación visual para el template (0 o valor)
        cant_entrada=Case(
            When(movimiento__almacen_destino=almacen, then=F('cantidad')),
            default=Value(0),
            output_field=DecimalField(max_digits=12, decimal_places=2)
        ),
        cant_salida=Case(
            When(movimiento__almacen_origen=almacen, then=F('cantidad')),
            default=Value(0),
            output_field=DecimalField(max_digits=12, decimal_places=2)
        )
    ).annotate(
        # Calcular Saldo Acumulado (Running Total) cronológicamente
        saldo_historico=Window(
            expression=Sum('flujo_cantidad'),
            order_by=[F('movimiento__fecha').asc(), F('id').asc()]
        )
    ).select_related('movimiento', 'movimiento__creado_por', 'requerimiento').order_by('-movimiento__fecha', '-id')

    movimientos_visuales = []
    
    # 2. Procesamiento ligero para etiquetas visuales
    for detalle in movimientos:
        mov = detalle.movimiento
        tipo_original = mov.tipo
        
        # Ajuste de etiqueta visual para transferencias
        tipo_visual = tipo_original
        if detalle.cant_entrada > 0 and 'SALIDA' in tipo_original:
             tipo_visual = 'TRANSFERENCIA_ENTRADA'

        detalle.tipo_visual = tipo_visual
        
        # Mapeo de campos anotados a los nombres que espera el template
        detalle.cantidad_entrada = detalle.cant_entrada
        detalle.cantidad_salida = detalle.cant_salida
        detalle.saldo_calculado = detalle.saldo_historico # Usamos el valor calculado por BD
        
        # Lógica de etiqueta de Asignación
        if detalle.es_stock_libre:
            detalle.asignacion_visual = "STOCK LIBRE"
        elif detalle.requerimiento:
            detalle.asignacion_visual = f"{detalle.requerimiento.codigo}"
        elif mov.requerimiento:
            detalle.asignacion_visual = f"{mov.requerimiento.codigo}"
        else:
            detalle.asignacion_visual = "STOCK LIBRE"

        movimientos_visuales.append(detalle)

    context = {
        'almacen': almacen,
        'material': material,
        'movimientos': movimientos_visuales
    }
    return render(request, 'logistica/kardex_producto.html', context)

def requerimiento_list(request):
    """
    Lista de seguimiento de Requerimientos (Pedidos de Obra).
    """
    requerimientos = Requerimiento.objects.select_related('proyecto', 'creado_por').order_by('-fecha_solicitud')
    
    context = {
        'requerimientos': requerimientos
    }
    return render(request, 'logistica/requerimiento_list.html', context)

def requerimiento_detail(request, req_id):
    """
    Ver el progreso de un requerimiento: Qué se pidió vs Qué se ha entregado.
    """
    req = get_object_or_404(Requerimiento, id=req_id)
    # Salidas confirmadas que han atendido este requerimiento
    salidas = req.salidas.filter(estado='CONFIRMADO').order_by('-fecha')
    
    context = {
        'req': req,
        'salidas': salidas
    }
    return render(request, 'logistica/requerimiento_detail.html', context)

def requerimiento_create(request):
    """
    Crea un nuevo Requerimiento (Pedido de Materiales).
    """
    if request.method == 'POST':
        form = RequerimientoForm(request.POST)
        formset = DetalleRequerimientoFormSet(request.POST)
        
        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    req = form.save(commit=False)
                    req.creado_por = request.user
                    
                    # Asignamos el proyecto por defecto (o el del usuario si existiera lógica)
                    # Aquí asumimos el primer proyecto activo para evitar errores
                    req.proyecto = Proyecto.objects.first() 
                    
                    req.save()
                    
                    formset.instance = req
                    formset.save()
                    
                    messages.success(request, f'Requerimiento {req.codigo} creado exitosamente.')
                    return redirect('requerimiento_list')
            except Exception as e:
                messages.error(request, f"Error al guardar: {e}")
    else:
        form = RequerimientoForm(initial={'fecha_solicitud': timezone.now().date()})
        formset = DetalleRequerimientoFormSet()

    context = {
        'form': form,
        'formset': formset,
        'titulo': "Nuevo Requerimiento",
        'boton_texto': "Crear Pedido",
        'materiales_disponibles': Material.objects.filter(activo=True).order_by('codigo'),
        'trabajadores_disponibles': Trabajador.objects.filter(activo=True).order_by('nombres'),
    }
    return render(request, 'logistica/requerimiento_form.html', context)

# ==========================================
# 3. CREACIÓN Y GESTIÓN DE MOVIMIENTOS
# ==========================================

def operacion_almacen(request, tipo_accion, almacen_id):
    """
    Vista principal para registrar Ingresos y Salidas.
    CORRECCIONES: Usuario, Proyecto, Filtros y GENERADOR DOBLE (NI/VS).
    """
    if str(almacen_id) == '00000000-0000-0000-0000-000000000000':
        # Si no viene almacén en la URL, intentamos usar el de la sesión
        if hasattr(request, 'almacen_activo') and request.almacen_activo:
            almacen = request.almacen_activo
        else:
            almacen = None
    else:
        almacen = get_object_or_404(Almacen, id=almacen_id)
    
    tipo_default = 'SALIDA_OBRA' if tipo_accion == 'salida' else 'INGRESO_COMPRA'
    if request.GET.get('tipo'):
        tipo_default = request.GET.get('tipo')

    # Lógica dinámica: Si estamos en un almacén, definimos si es Origen o Destino
    initial_data = {
        'tipo': tipo_default
    }
    if almacen:
        if tipo_accion == 'ingreso':
            initial_data['almacen_destino'] = almacen
        else:
            initial_data['almacen_origen'] = almacen
    
    # Determinamos el ID del almacén para filtrar los activos en el formset
    filtro_almacen_id = almacen.id if almacen else None

    if request.method == 'POST':
        form = MovimientoForm(request.POST, tipo_accion=tipo_accion)
        # Si es POST y no tenemos almacén en URL (vista genérica), intentamos sacarlo del POST
        if not filtro_almacen_id:
            filtro_almacen_id = request.POST.get('almacen_origen')
            
        # Obtenemos el tipo real del POST para pasarlo al formset (para filtrar activos en REINGRESO_LIMA)
        tipo_seleccionado = request.POST.get('tipo', tipo_default)
        formset = DetalleMovimientoFormSet(request.POST, form_kwargs={'tipo_accion': tipo_accion, 'almacen_id': filtro_almacen_id, 'tipo_movimiento': tipo_seleccionado})
        if form.is_valid():
            nuevo_mov = form.save(commit=False)
            
            # 1. ASIGNAR USUARIO
            nuevo_mov.creado_por = request.user 
            
            # 2. ASIGNAR PROYECTO
            if almacen and almacen.proyecto:
                nuevo_mov.proyecto = almacen.proyecto
            else:
                # Fallback por si el almacén no tiene proyecto
                # Intentamos obtener el proyecto del almacén seleccionado en el formulario
                almacen_seleccionado = nuevo_mov.almacen_origen or nuevo_mov.almacen_destino
                nuevo_mov.proyecto = almacen_seleccionado.proyecto if almacen_seleccionado else Proyecto.objects.first()

            # 3. GENERADOR DE CÓDIGO AUTOMÁTICO (NI o VS)
            # Solo generamos si no tiene uno manual
            if not nuevo_mov.nota_ingreso:
                es_salida = 'SALIDA' in nuevo_mov.tipo or 'CONSUMO' in nuevo_mov.tipo
                es_ingreso = 'INGRESO' in nuevo_mov.tipo or 'DEVOLUCION' in nuevo_mov.tipo or 'ENTRADA' in nuevo_mov.tipo
                
                prefix = None
                if es_salida:
                    prefix = 'VS-'
                elif es_ingreso:
                    prefix = 'NI-'
                
                if prefix:
                    # Buscamos el último código de ese tipo (VS o NI)
                    ultimo = Movimiento.objects.filter(nota_ingreso__startswith=prefix).order_by('-nota_ingreso').first()
                    
                    contador = 1
                    if ultimo:
                        try:
                            # Ejemplo: NI-0007 -> toma "0007" -> suma 1 -> 8
                            contador = int(ultimo.nota_ingreso.split('-')[1]) + 1
                        except:
                            contador = 1
                    
                    # Asignamos: NI-0008 o VS-0003
                    nuevo_mov.nota_ingreso = f"{prefix}{contador:04d}"

            # --- GUARDAR ---
            nuevo_mov.save()
            form.save_m2m() 
            
            formset = DetalleMovimientoFormSet(request.POST, instance=nuevo_mov, form_kwargs={'tipo_accion': tipo_accion, 'almacen_id': filtro_almacen_id, 'tipo_movimiento': tipo_seleccionado})
            if formset.is_valid():
                formset.save()
                
                # RESPUESTA AJAX (JSON) PARA TOAST
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'message': f'Operación {nuevo_mov.nota_ingreso} registrada exitosamente.',
                        'redirect_url': reverse('movimiento_list')
                    })

                messages.success(request, f'Operación {nuevo_mov.nota_ingreso} registrada exitosamente.')
                return redirect('movimiento_list')
            else:
                nuevo_mov.delete()
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': 'Error en los detalles de los materiales.'})
                messages.error(request, 'Error en los detalles de los materiales.')
    else:
        form = MovimientoForm(initial=initial_data, tipo_accion=tipo_accion)
        formset = DetalleMovimientoFormSet(form_kwargs={'tipo_accion': tipo_accion, 'almacen_id': filtro_almacen_id, 'tipo_movimiento': tipo_default})

    # =========================================================================
    # 🛡️ FILTRO DINÁMICO DEL DESPLEGABLE
    # =========================================================================
    choices_originales = form.fields['tipo'].choices
    choices_filtradas = []

    for key, label in choices_originales:
        if not key:
            choices_filtradas.append((key, label))
            continue
        clave = str(key).upper()
        if tipo_accion == 'ingreso':
            # Excluimos DEVOLUCION_LIMA porque es una salida (aunque diga DEVOLUCION)
            if ('INGRESO' in clave or 'DEVOLUCION' in clave or 'ENTRADA' in clave) and clave != 'DEVOLUCION_LIMA':
                choices_filtradas.append((key, label))
        elif tipo_accion == 'salida':
            if 'SALIDA' in clave or 'CONSUMO' in clave or clave == 'DEVOLUCION_LIMA':
                choices_filtradas.append((key, label))

    form.fields['tipo'].choices = choices_filtradas
    # =========================================================================

    # =========================================================================
    # 🛡️ VALIDACIÓN FRONTEND: MAPA DE MATERIALES POR REQUERIMIENTO
    # =========================================================================
    reqs_map = {}
    mats_reqs_map = {} # Nuevo: Mapa inverso (Material -> Requerimientos)
    # Obtenemos el queryset que ya filtró el formulario (solo pendientes/parciales)
    if 'requerimiento' in form.fields:
        qs_reqs = form.fields['requerimiento'].queryset
        # Obtenemos los materiales válidos para cada requerimiento en una sola consulta
        detalles_data = DetalleRequerimiento.objects.filter(
            requerimiento__in=qs_reqs
        ).values('requerimiento_id', 'material_id')
        
        for item in detalles_data:
            r_id = str(item['requerimiento_id'])
            m_id = str(item['material_id'])
            if r_id not in reqs_map: reqs_map[r_id] = []
            reqs_map[r_id].append(m_id)
            
            # Llenamos el mapa inverso
            if m_id not in mats_reqs_map: mats_reqs_map[m_id] = []
            mats_reqs_map[m_id].append(r_id)

    context = {
        'form': form,
        'formset': formset,
        'almacen': almacen,
        'titulo': f"{'Salida' if tipo_accion == 'salida' else 'Ingreso'} de Materiales{' - ' + almacen.nombre if almacen else ''}",
        'boton_texto': f"Confirmar {'Salida' if tipo_accion == 'salida' else 'Ingreso'}",
        'materiales_disponibles': Material.objects.filter(activo=True).order_by('codigo'),
        'reqs_materiales_json': json.dumps(reqs_map),
        'mats_reqs_json': json.dumps(mats_reqs_map), # Enviamos el nuevo mapa al template
        'materiales_tipos_json': json.dumps({str(id): tipo for id, tipo in Material.objects.values_list('id', 'tipo')}), # Mapa de tipos para JS
    }
    return render(request, 'logistica/operacion_form.html', context)

def editar_movimiento(request, movimiento_id):
    """
    Permite editar un movimiento existente (Solo Borradores).
    """
    movimiento = get_object_or_404(Movimiento, id=movimiento_id)
    
    if movimiento.estado != 'BORRADOR':
        messages.error(request, "No se puede editar un movimiento confirmado.")
        return redirect('movimiento_list')

    # Determinamos el tipo de acción basado en el movimiento existente
    es_ingreso = 'INGRESO' in movimiento.tipo or 'DEVOLUCION' in movimiento.tipo or 'ENTRADA' in movimiento.tipo
    tipo_accion = 'ingreso' if es_ingreso else 'salida'
    
    filtro_almacen_id = movimiento.almacen_origen_id

    if request.method == 'POST':
        form = MovimientoForm(request.POST, instance=movimiento, tipo_accion=tipo_accion)
        formset = DetalleMovimientoFormSet(request.POST, instance=movimiento, form_kwargs={'tipo_accion': tipo_accion, 'almacen_id': filtro_almacen_id, 'tipo_movimiento': movimiento.tipo})
        
        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    form.save()
                    formset.save()
                    messages.success(request, 'Movimiento actualizado correctamente.')
                    return redirect('movimiento_list')
            except Exception as e:
                messages.error(request, f"Error al actualizar: {e}")
    else:
        form = MovimientoForm(instance=movimiento, tipo_accion=tipo_accion)
        formset = DetalleMovimientoFormSet(instance=movimiento, form_kwargs={'tipo_accion': tipo_accion, 'almacen_id': filtro_almacen_id, 'tipo_movimiento': movimiento.tipo})

    almacen_contexto = movimiento.almacen_origen or movimiento.almacen_destino

    context = {
        'form': form,
        'formset': formset,
        'titulo': f"Editar {movimiento.get_tipo_display()}",
        'boton_texto': "Guardar Cambios",
        'almacen': almacen_contexto,
    }
    return render(request, 'logistica/operacion_form.html', context)

# ==========================================
# 4. ACCIONES DE MOVIMIENTO (CONFIRMAR / ELIMINAR)
# ==========================================

def confirmar_movimiento_web(request, movimiento_id):
    movimiento = get_object_or_404(Movimiento, id=movimiento_id)
    
    if movimiento.estado == 'BORRADOR':
        try:
            # --- RED DE SEGURIDAD (Generador Tardío de Código) ---
            if not movimiento.nota_ingreso or movimiento.nota_ingreso == 'Por Generar':
                es_salida = 'SALIDA' in movimiento.tipo or 'CONSUMO' in movimiento.tipo
                es_ingreso = 'INGRESO' in movimiento.tipo or 'DEVOLUCION' in movimiento.tipo or 'ENTRADA' in movimiento.tipo
                prefix = 'VS-' if es_salida else ('NI-' if es_ingreso else None)
                
                if prefix:
                    ultimo = Movimiento.objects.filter(nota_ingreso__startswith=prefix).order_by('-nota_ingreso').first()
                    contador = 1
                    if ultimo:
                        try:
                            contador = int(ultimo.nota_ingreso.split('-')[1]) + 1
                        except:
                            contador = 1
                    movimiento.nota_ingreso = f"{prefix}{contador:04d}"
                movimiento.save()
            
            # --- USAMOS EL SERVICIO CENTRALIZADO ---
            # Esto actualiza Stock y Existencia correctamente
            KardexService.confirmar_movimiento(movimiento.id)
            messages.success(request, f'Movimiento {movimiento.nota_ingreso} confirmado correctamente.')

        except ValidationError as e:
            messages.error(request, f'Validación: {e.message}')
        except Exception as e:
            messages.error(request, f'Error al procesar: {str(e)}')
            
    else:
        messages.warning(request, 'Este movimiento ya estaba confirmado.')
    
    return redirect('movimiento_list')

def anular_movimiento(request, movimiento_id):
    """
    Anula un movimiento. Si es Borrador lo cancela, si es Confirmado revierte stock.
    """
    movimiento = get_object_or_404(Movimiento, id=movimiento_id)
    
    try:
        KardexService.anular_movimiento(movimiento.id)
        messages.success(request, f'Movimiento {movimiento.nota_ingreso} ANULADO correctamente. Stock revertido.')
    except ValidationError as e:
        messages.error(request, f"No se pudo anular: {e.message}")
        
    return redirect('movimiento_list')

# ==========================================
# 5. APIs Y HERRAMIENTAS (SCRIPTS)
# ==========================================

def api_consultar_stock(request, almacen_id, material_id):
    """
    API JSON: Devuelve el stock actual de un material y su costo promedio.
    """
    try:
        if str(almacen_id) == '00000000-0000-0000-0000-000000000000':
            return JsonResponse({'stock': 0, 'precio': 0})

        item = Stock.objects.filter(almacen_id=almacen_id, material_id=material_id).first()
        stock_actual = item.cantidad if item else 0
        
        # Obtener Costo Promedio (Existencia del Proyecto)
        precio = 0
        almacen = item.almacen if item else Almacen.objects.filter(id=almacen_id).first()
        if almacen and almacen.proyecto:
            existencia = Existencia.objects.filter(proyecto=almacen.proyecto, material_id=material_id).first()
            if existencia:
                precio = existencia.costo_promedio

        return JsonResponse({'stock': stock_actual, 'precio': precio})
    except Exception as e:
        return JsonResponse({'stock': 0, 'precio': 0, 'error': str(e)})

def api_crear_trabajador(request):
    """
    API para creación rápida de trabajadores desde el formulario de salida.
    """
    if request.method == 'POST':
        nombres = request.POST.get('nombres')
        apellidos = request.POST.get('apellidos')
        dni = request.POST.get('dni')
        
        if not all([nombres, apellidos, dni]):
            return JsonResponse({'success': False, 'error': 'Faltan datos obligatorios (Nombre, Apellido, DNI).'})
        
        if not dni.isdigit() or len(dni) != 8:
            return JsonResponse({'success': False, 'error': 'El DNI debe tener exactamente 8 dígitos numéricos.'})

        if Trabajador.objects.filter(dni=dni).exists():
            return JsonResponse({'success': False, 'error': f'Ya existe un trabajador con DNI {dni}.'})
        
        try:
            t = Trabajador.objects.create(
                nombres=nombres.upper(),
                apellidos=apellidos.upper(),
                dni=dni,
                activo=True
            )
            return JsonResponse({'success': True, 'id': t.id, 'text': str(t)})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

def api_buscar_trabajador(request):
    """
    Buscador optimizado para miles de registros.
    """
    q = request.GET.get('q', '').strip()
    if len(q) < 2:
        return JsonResponse({'results': []})
    
    qs = Trabajador.objects.filter(activo=True).filter(
        Q(nombres__icontains=q) | Q(apellidos__icontains=q) | Q(dni__icontains=q)
    ).values('id', 'nombres', 'apellidos', 'dni')[:20] # Limitamos a 20 resultados
    
    return JsonResponse({'results': list(qs)})

# ==========================================
# 6. ZONA DE PELIGRO (ADMINISTRACIÓN)
# ==========================================

def reset_database(request):
    """
    Limpia toda la data transaccional (Movimientos, Stock, Requerimientos).
    Mantiene catálogos (Materiales, Proyectos, Usuarios).
    Requiere ser Superusuario y confirmar contraseña.
    """
    if not request.user.is_superuser:
        messages.error(request, "Acceso denegado. Solo administradores.")
        return redirect('dashboard')

    # Validar que estemos en modo global (sin almacén activo)
    if getattr(request, 'almacen_activo', None):
        messages.warning(request, "Para resetear el sistema debes estar en 'Vista Global'. Por seguridad, sal del almacén actual.")
        return redirect('dashboard')

    if request.method == 'POST':
        password = request.POST.get('password')
        if not request.user.check_password(password):
            messages.error(request, "Contraseña incorrecta. No se realizaron cambios.")
        else:
            try:
                with transaction.atomic():
                    # 1. Eliminar detalles (Rompe dependencia con Activos)
                    DetalleMovimiento.objects.all().delete()
                    
                    # 2. Eliminar Activos Fijos y Asignaciones (Rompe dependencia con Movimientos)
                    AsignacionActivo.objects.all().delete()
                    Activo.objects.all().delete()
                    Kit.objects.all().delete()

                    # 3. Eliminar Movimientos (Ahora sí se puede)
                    Movimiento.objects.all().delete()
                    
                    # 4. Eliminar Stock y Costos
                    Stock.objects.all().delete()
                    Existencia.objects.all().delete()
                    
                    # 5. Eliminar Requerimientos
                    DetalleRequerimiento.objects.all().delete()
                    Requerimiento.objects.all().delete()
                
                messages.success(request, "✅ Base de datos operativa reiniciada correctamente.")
                return redirect('dashboard')
            except Exception as e:
                messages.error(request, f"Error al limpiar datos: {e}")

    return render(request, 'logistica/reset_db.html')

def cerrar_requerimiento(request, req_id):
    """
    Cierra forzosamente un requerimiento (ej: ya no se necesita el saldo pendiente).
    """
    req = get_object_or_404(Requerimiento, id=req_id)
    
    if req.estado not in ['TOTAL', 'CANCELADO']:
        # Verificar si hay stock ingresado que no se entregó (Sobrante)
        sobrantes = []
        for det in req.detalles.all():
            remanente = det.cantidad_ingresada - det.cantidad_atendida
            if remanente > 0:
                sobrantes.append(f"{det.material.codigo} ({remanente})")

        req.estado = 'TOTAL' # Lo marcamos como completado
        req.observacion += "\n[SISTEMA] Cerrado manualmente por el usuario (Saldo anulado)."
        req.save()
        
        if sobrantes:
            lista = ", ".join(sobrantes)
            messages.warning(request, f"⚠️ ATENCIÓN: El requerimiento {req.codigo} se cerró con materiales ingresados NO entregados: {lista}. Estos pasan a ser STOCK LIBRE.")
        else:
            messages.success(request, f"Requerimiento {req.codigo} finalizado manualmente.")
        
    return redirect('requerimiento_detail', req_id=req.id)

# ==========================================
# 7. EXPORTACIÓN A EXCEL
# ==========================================

def exportar_inventario_excel(request):
    """
    Genera un Excel con el stock actual filtrado por la búsqueda.
    """
    query = request.GET.get('q')
    stocks_filter = Stock.objects.select_related('material', 'material__categoria', 'almacen')

    if query:
        stocks_filter = stocks_filter.filter(
            Q(material__codigo__icontains=query) |
            Q(material__descripcion__icontains=query)
        )
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario Físico"

    # Encabezados
    headers = ["Almacén", "Código", "Material", "Categoría", "Unidad", "Stock Actual", "Mínimo", "Ubicación"]
    ws.append(headers)
    
    # Estilo Encabezado
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Datos
    for stock in stocks_filter:
        ws.append([
            stock.almacen.nombre,
            stock.material.codigo,
            stock.material.descripcion,
            stock.material.categoria.nombre if stock.material.categoria else '-',
            stock.material.unidad_medida,
            stock.cantidad,
            stock.cantidad_minima,
            stock.ubicacion_pasillo
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Inventario_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

def exportar_kardex_excel(request, almacen_id, material_id):
    """
    Genera el Kardex de un producto específico en Excel.
    """
    almacen = get_object_or_404(Almacen, id=almacen_id)
    material = get_object_or_404(Material, id=material_id)

    # Reutilizamos la lógica de cálculo de saldos
    stock_item = Stock.objects.filter(almacen=almacen, material=material).first()
    saldo_actual = stock_item.cantidad if stock_item else Decimal(0)

    movimientos = DetalleMovimiento.objects.filter(
        material=material,
        movimiento__estado='CONFIRMADO'
    ).filter(
        Q(movimiento__almacen_origen=almacen) | 
        Q(movimiento__almacen_destino=almacen)
    ).select_related('movimiento', 'movimiento__creado_por').order_by('-movimiento__fecha')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Kardex {material.codigo}"

    # Información General
    ws.append(["Material:", f"{material.codigo} - {material.descripcion}"])
    ws.append(["Almacén:", almacen.nombre])
    ws.append([]) # Espacio

    # Encabezados Tabla
    headers = ["Fecha", "Tipo Operación", "Documento", "Asignación / Destino", "Entrada", "Salida", "Saldo", "Usuario"]
    ws.append(headers)
    
    # Estilo Encabezado
    for cell in ws[4]: # Fila 4 porque agregamos 3 filas antes
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")

    # Procesar datos (Misma lógica que la vista web)
    rows = []
    for detalle in movimientos:
        mov = detalle.movimiento
        
        # Determinar si es entrada o salida para este almacén
        id_actual = str(almacen.id)
        id_destino = str(mov.almacen_destino_id) if mov.almacen_destino_id else ''
        id_origen = str(mov.almacen_origen_id) if mov.almacen_origen_id else ''

        soy_destino = (id_destino == id_actual)
        soy_origen = (id_origen == id_actual)

        es_ingreso = False
        es_salida = False

        if soy_destino:
            es_ingreso = True
        elif soy_origen:
            es_salida = True
        else:
            # Fallback
            if 'INGRESO' in mov.tipo or 'DEVOLUCION' in mov.tipo or 'ENTRADA' in mov.tipo: es_ingreso = True
            elif 'SALIDA' in mov.tipo or 'CONSUMO' in mov.tipo: es_salida = True

        entrada = detalle.cantidad if es_ingreso else 0
        salida = detalle.cantidad if es_salida else 0
        
        # Lógica de visualización de asignación
        asignacion_str = "STOCK LIBRE"
        if detalle.es_stock_libre:
            asignacion_str = "STOCK LIBRE"
        elif detalle.requerimiento:
            asignacion_str = detalle.requerimiento.codigo
        elif mov.requerimiento:
            asignacion_str = mov.requerimiento.codigo

        # Guardamos en lista temporal para invertir el orden si quisiéramos, 
        # pero aquí lo agregamos directo y calculamos saldo hacia atrás
        rows.append([
            mov.fecha.strftime("%d/%m/%Y %H:%M"),
            mov.get_tipo_display(),
            f"{mov.nota_ingreso or ''} {mov.documento_referencia or ''}",
            asignacion_str,
            entrada,
            salida,
            saldo_actual,
            mov.creado_por.username
        ])

        # Recalcular saldo anterior
        if es_ingreso: saldo_actual -= detalle.cantidad
        elif es_salida: saldo_actual += detalle.cantidad

    # Escribir filas
    for row in rows:
        ws.append(row)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Kardex_{material.codigo}.xlsx"'
    wb.save(response)
    return response

def exportar_activos_externos_excel(request):
    """
    Genera un reporte Excel de los activos que están actualmente en Sede Central (Devueltos).
    Busca el último movimiento de salida 'DEVOLUCION_LIMA' para obtener la fecha y guía.
    """
    # Filtramos activos que ya no están en obra
    activos_externos = Activo.objects.filter(estado='DEVUELTO_EXTERNO').order_by('codigo')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Activos en Sede Central"

    # Encabezados
    headers = ["Código", "Activo", "Marca", "Modelo", "Serie", "Fecha Devolución", "Guía Remisión / Ref.", "Usuario Devolvió"]
    ws.append(headers)

    # Estilo Encabezado (Rojo para diferenciar que son externos)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid") 
        cell.alignment = Alignment(horizontal="center")

    for activo in activos_externos:
        # Buscamos el movimiento que causó esta devolución (El último DEVOLUCION_LIMA)
        ultimo_mov = DetalleMovimiento.objects.filter(
            activo=activo,
            movimiento__tipo='DEVOLUCION_LIMA'
        ).select_related('movimiento', 'movimiento__creado_por').order_by('-movimiento__fecha').first()

        fecha_dev = "S/D"
        guia = "-"
        usuario = "-"

        if ultimo_mov:
            fecha_dev = ultimo_mov.movimiento.fecha.strftime("%d/%m/%Y %H:%M")
            guia = f"{ultimo_mov.movimiento.nota_ingreso} / {ultimo_mov.movimiento.documento_referencia}"
            usuario = ultimo_mov.movimiento.creado_por.username

        ws.append([
            activo.codigo,
            activo.nombre,
            activo.marca,
            activo.modelo,
            activo.serie,
            fecha_dev,
            guia,
            usuario
        ])

    # Ajuste visual de columnas
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 25

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Activos_SedeCentral_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

# ==========================================
# 8. CARGA MASIVA DE DATOS
# ==========================================

def descargar_plantilla_importacion(request):
    """
    Genera y descarga una plantilla Excel vacía con las cabeceras correctas.
    """
    wb = openpyxl.Workbook()
    
    # --- HOJA 1: MATERIALES ---
    ws1 = wb.active
    ws1.title = "Materiales"
    headers_mat = ["CODIGO", "DESCRIPCION", "UNIDAD", "CATEGORIA", "TIPO (CONSUMIBLE/ACTIVO_FIJO/EPP)"]
    ws1.append(headers_mat)
    # Ejemplo
    ws1.append(["CEM-001", "CEMENTO SOL TIPO I", "BOL", "ALBAÑILERIA", "CONSUMIBLE"])
    
    # Estilo
    for cell in ws1[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")

    # --- HOJA 2: ACTIVOS ---
    ws2 = wb.create_sheet("Activos")
    headers_act = ["CODIGO_ACTIVO", "SERIE", "NOMBRE", "MARCA", "MODELO", "COD_MATERIAL_CATALOGO", "ALMACEN_UBICACION", "VALOR_COMPRA"]
    ws2.append(headers_act)
    # Ejemplo
    ws2.append(["TAL-001", "SN-998877", "TALADRO PERCUTOR", "HILTI", "TE-30", "TAL-HIL", "Almacén Central", 500.00])

    # Estilo
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")

    # --- HOJA 3: STOCK INICIAL ---
    ws3 = wb.create_sheet("StockInicial")
    headers_stock = ["CODIGO_MATERIAL", "CANTIDAD", "COSTO_UNITARIO", "ALMACEN_DESTINO"]
    ws3.append(headers_stock)
    # Ejemplo
    ws3.append(["CEM-001", 500, 22.50, "Almacén Central"])

    # Estilo
    for cell in ws3[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid")

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Plantilla_Carga_Masiva.xlsx"'
    wb.save(response)
    return response

def importar_datos_excel(request):
    """
    Procesa el archivo Excel subido y crea/actualiza registros.
    """
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, "Acceso denegado. Solo administradores pueden importar datos.")
        return redirect('dashboard')

    if request.method == 'POST':
        form = ImportarDatosForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['archivo_excel']
            try:
                wb = openpyxl.load_workbook(excel_file, data_only=True)
                count_mat = 0
                count_act = 0
                count_stock = 0
                errores = []
                resumen = {}
                
                # 1. Importar Materiales
                if 'Materiales' in wb.sheetnames:
                    sheet = wb['Materiales']
                    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                        try:
                            row_data = list(row) + [None] * (5 - len(row))
                            codigo, descripcion, unidad, cat_nombre, tipo = row_data[:5]
                            
                            if not codigo: continue
                            
                            with transaction.atomic():
                                categoria = None
                                if cat_nombre:
                                    nombre_cat = str(cat_nombre).strip().upper()
                                    # 1. Buscamos por nombre exacto
                                    categoria = Categoria.objects.filter(nombre=nombre_cat).first()
                                    
                                    if not categoria:
                                        # 2. Si no existe por nombre, buscamos por CÓDIGO (Inferencia)
                                        # Si el código ya existe (ej: ACT), asumimos que es la misma categoría
                                        # y la reutilizamos para evitar duplicados o errores.
                                        base_code = nombre_cat[:3].upper()
                                        categoria = Categoria.objects.filter(codigo=base_code).first()
                                        
                                        if not categoria:
                                            # 3. Solo si no existe ni nombre ni código, CREAMOS
                                            categoria = Categoria.objects.create(nombre=nombre_cat, codigo=base_code)
                                
                                Material.objects.update_or_create(
                                    codigo=str(codigo).strip().upper(),
                                    defaults={
                                        'descripcion': str(descripcion).strip().upper() if descripcion else 'SIN DESCRIPCION',
                                        'unidad_medida': str(unidad).strip().upper() if unidad else 'UND',
                                        'categoria': categoria,
                                        'tipo': str(tipo).strip().upper() if tipo else 'CONSUMIBLE',
                                        'activo': True
                                    }
                                )
                                count_mat += 1
                        except Exception as e:
                            errores.append(f"[Materiales] Fila {i} (Código: {row[0] if row else '?'}): {str(e)}")

                # 2. Importar Activos
                if 'Activos' in wb.sheetnames:
                    sheet = wb['Activos']
                    activos_nuevos_por_almacen = {} # Diccionario para agrupar activos nuevos por almacén
                    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                        try:
                            row_data = list(row) + [None] * (8 - len(row))
                            cod_activo, serie, nombre, marca, modelo, cod_material, nombre_almacen, valor_compra = row_data[:8]
                            
                            if not cod_activo: continue
                            
                            with transaction.atomic():
                                material = None
                                if cod_material:
                                    material = Material.objects.filter(codigo=str(cod_material).strip().upper()).first()
                                    if not material:
                                        raise ValueError(f"Material catálogo '{cod_material}' no existe")
                                
                                ubicacion = None
                                if nombre_almacen:
                                    ubicacion = Almacen.objects.filter(nombre__icontains=str(nombre_almacen).strip()).first()
                                    if not ubicacion:
                                        raise ValueError(f"Almacén '{nombre_almacen}' no encontrado")
                                
                                # Parsear Costo (Valor de Compra)
                                try:
                                    costo = Decimal(str(valor_compra)) if valor_compra else Decimal(0)
                                except:
                                    costo = Decimal(0)

                                activo, created = Activo.objects.update_or_create(
                                    codigo=str(cod_activo).strip().upper(),
                                    defaults={
                                        'serie': str(serie).strip().upper() if serie else '',
                                        'nombre': str(nombre).strip().upper() if nombre else 'ACTIVO SIN NOMBRE',
                                        'marca': str(marca).strip().upper() if marca else '',
                                        'modelo': str(modelo).strip().upper() if modelo else '',
                                        'material': material,
                                        'ubicacion': ubicacion,
                                        'estado': 'DISPONIBLE',
                                        'valor_compra': costo,
                                    }
                                )
                                count_act += 1
                                
                                # --- AUTOMATIZACIÓN DE STOCK ---
                                # Si el activo es NUEVO y tiene ubicación/material, lo encolamos para crearle stock
                                if created and material and ubicacion:
                                    alm_id = ubicacion.id
                                    if alm_id not in activos_nuevos_por_almacen:
                                        activos_nuevos_por_almacen[alm_id] = {
                                            'almacen': ubicacion,
                                            'items': []
                                        }
                                    activos_nuevos_por_almacen[alm_id]['items'].append({
                                        'activo': activo,
                                        'material': material,
                                        'costo': costo
                                    })

                        except Exception as e:
                            errores.append(f"[Activos] Fila {i} (Activo: {row[0] if row else '?'}): {str(e)}")

                    # --- PROCESAR STOCK AUTOMÁTICO PARA ACTIVOS NUEVOS ---
                    for alm_id, data in activos_nuevos_por_almacen.items():
                        try:
                            with transaction.atomic():
                                almacen_obj = data['almacen']
                                items = data['items']
                                
                                # Crear Movimiento de Ajuste (Ingreso)
                                mov = Movimiento.objects.create(
                                    proyecto=almacen_obj.proyecto,
                                    tipo='INGRESO_COMPRA', # Cambiado para generar NI y consistencia
                                    almacen_destino=almacen_obj,
                                    fecha=timezone.now(),
                                    creado_por=request.user,
                                    documento_referencia='CARGA_MASIVA_ACT',
                                    observacion='Generación automática de stock por carga masiva de activos',
                                    estado='BORRADOR'
                                )
                                
                                for item in items:
                                    # Validar costo para evitar error de Kardex (Requiere > 0)
                                    costo_final = item['costo']
                                    if costo_final <= 0:
                                        costo_final = Decimal('1.00') # Valor nominal por defecto

                                    DetalleMovimiento.objects.create(
                                        movimiento=mov,
                                        material=item['material'],
                                        cantidad=1,
                                        costo_unitario=costo_final,
                                        activo=item['activo'],
                                        es_stock_libre=True
                                    )
                                
                                KardexService.confirmar_movimiento(mov.id)
                                count_stock += len(items) # Sumamos al contador de stock procesado
                        except Exception as e:
                            errores.append(f"[AutoStock] Error generando stock en {data['almacen'].nombre}: {str(e)}")

                # 3. Importar Stock Inicial
                if 'StockInicial' in wb.sheetnames:
                    sheet = wb['StockInicial']
                    batch_movimientos = {}
                    
                    # Fase 1: Lectura y Validación
                    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                        try:
                            row_data = list(row) + [None] * (4 - len(row))
                            cod_mat, cant, costo, nom_almacen = row_data[:4]
                            
                            if not cod_mat or not cant or not nom_almacen: continue
                            
                            # Validar Material
                            material = Material.objects.filter(codigo=str(cod_mat).strip().upper()).first()
                            if not material: 
                                raise ValueError(f"Material '{cod_mat}' no existe")
                            
                            nom_almacen = str(nom_almacen).strip()
                            if nom_almacen not in batch_movimientos:
                                batch_movimientos[nom_almacen] = []
                            
                            batch_movimientos[nom_almacen].append({
                                'material': material,
                                'cantidad': Decimal(str(cant)),
                                'costo': Decimal(str(costo)) if costo else Decimal(0),
                                'fila': i
                            })
                        except Exception as e:
                            errores.append(f"[StockInicial] Fila {i}: {str(e)}")

                    # Fase 2: Procesamiento por Lotes (Almacén)
                    for nombre_almacen, items in batch_movimientos.items():
                        try:
                            with transaction.atomic():
                                almacen = Almacen.objects.filter(nombre__icontains=nombre_almacen).first()
                                if not almacen: 
                                    raise ValueError(f"Almacén '{nombre_almacen}' no encontrado en BD")
                                
                                mov = Movimiento.objects.create(
                                    proyecto=almacen.proyecto,
                                    tipo='INGRESO_COMPRA', # Cambiamos a Ingreso para que genere NI y sea explícito
                                    almacen_destino=almacen,
                                    fecha=timezone.now(),
                                    creado_por=request.user,
                                    documento_referencia='CARGA_MASIVA',
                                    observacion='Carga Inicial de Stock desde Excel',
                                    estado='BORRADOR'
                                )
                                
                                temp_count = 0
                                for item in items:
                                    DetalleMovimiento.objects.create(
                                        movimiento=mov,
                                        material=item['material'],
                                        cantidad=item['cantidad'],
                                        costo_unitario=item['costo'],
                                        es_stock_libre=True
                                    )
                                    temp_count += 1
                                
                                KardexService.confirmar_movimiento(mov.id)
                                count_stock += temp_count
                        except Exception as e:
                            filas_afectadas = ", ".join([str(x['fila']) for x in items])
                            errores.append(f"[StockInicial] Error procesando lote '{nombre_almacen}' (Filas {filas_afectadas}): {str(e)}")
                
                resumen = {
                    'materiales': count_mat,
                    'activos': count_act,
                    'stock': count_stock
                }
                
                if errores:
                    messages.warning(request, f"Carga finalizada con {len(errores)} errores. Revise el reporte abajo.")
                else:
                    messages.success(request, f"✅ Carga Exitosa: {count_mat} Materiales, {count_act} Activos y {count_stock} líneas de Stock.")
                
                # Renderizamos con el contexto de errores (sin redirect para no perder la lista)
                context = {
                    'form': form,
                    'titulo': 'Carga Masiva de Datos (Catálogo y Activos)',
                    'errores': errores,
                    'resumen': resumen
                }
                return render(request, 'logistica/importar_datos.html', context)

            except Exception as e:
                messages.error(request, f"❌ Error al procesar el archivo: {e}")
    else:
        form = ImportarDatosForm()
    
    context = {
        'form': form,
        'titulo': 'Carga Masiva de Datos (Catálogo y Activos)'
    }
    return render(request, 'logistica/importar_datos.html', context)